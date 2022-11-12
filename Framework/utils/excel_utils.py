import openpyxl


def get_row_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.max_row


# print(get_row_count("E:\\BITM_Online_15\\Projects\\TestAutomationBITM15\\Framework\\data\\test_data.xlsx", "Sheet1"))


def get_column_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.max_column


# print(get_column_count("E:\\BITM_Online_15\\Projects\\TestAutomationBITM15\\Framework\\data\\test_data.xlsx", "Sheet1"))


def read_data(file, sheet, reading_row, reading_column):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.cell(row=reading_row, column=reading_column).value


# print(read_data("E:\\BITM_Online_15\\Projects\\TestAutomationBITM15\\Framework\\data\\test_data.xlsx", "Sheet1", 3, 1))


def write_data(file, sheet, reading_row, reading_column, data):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    worksheet.cell(row=reading_row, column=reading_column).value = data
    workbook.save(file)
